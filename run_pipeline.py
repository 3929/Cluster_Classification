import pandas as pd
import numpy as np
from FINAL import CLUSTER
import traceback
import warnings
import logging
import openpyxl

# Configure logging
logging.basicConfig(level=logging.DEBUG, 
                    format='%(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('clustering_debug.log'),
                        logging.StreamHandler()
                    ])
logger = logging.getLogger(__name__)

warnings.filterwarnings('ignore')

def main():
    # Initialize the CLUSTER class
    cluster = CLUSTER()
    
    # Load test dataset
    data_path = "CM1_RTOData.xlsx"
    
    # Configuration for the pipeline
    config = {
        'data_path': data_path,
        'quality_parameter': 'Blaine',
        'features': [
            'Blaine',
            'Total Feed',
            'RPHopperLevel',
            'MaterialTemp',
            'BMSepLoad',
            'MillRejects',
            'Residue'
        ],
        'filter_conditions': {
            'Blaine': {'min': 200, 'max': 350},
            'Total Feed': {'min': 100},
            'MillRejects': {'min': 100}
        }
    }
    
    try:
        # Load DataFrame and print column information
        df = pd.read_excel(data_path)
        print("DataFrame Columns:")
        print(df.columns.tolist())
        print("\nColumn Types:")
        print(df.dtypes)
        
        # Check if all features exist in the DataFrame
        missing_features = [feat for feat in config['features'] if feat not in df.columns]
        if missing_features:
            print(f"ERROR: Missing features: {missing_features}")
            return
        
        # Check data types of features
        print("\nFeature Details:")
        for feature in config['features']:
            print(f"\n{feature}:")
            print(f"  Data Type: {df[feature].dtype}")
            print(f"  Null Values: {df[feature].isnull().sum()}")
            print(f"  Value Range: {df[feature].min()} - {df[feature].max()}")
            
            # Convert to numeric if not already
            if df[feature].dtype == 'object':
                try:
                    df[feature] = pd.to_numeric(df[feature], errors='raise')
                    print(f"  Converted to numeric successfully")
                except:
                    print(f"  WARNING: Could not convert {feature} to numeric")

        results = cluster.run_clustering_pipeline(config)
        
        print("\nPipeline Results Summary:")
        print("------------------------")
        
        # Add safe access to results
        print(f"Raw records processed: {results.get('raw_records', 'No records processed')}")
        print(f"Hourly records generated: {results.get('hourly_records', 'No hourly records')}")
        print(f"Number of silos: {results.get('silo_count', 'No silos')}")
        print(f"Silos after processing: {results.get('processed_silo_count', 'No processed silos')}")
        print(f"Silos after filtering: {results.get('filtered_silo_count', 'No filtered silos')}")
        
        print("\nClustering Results by Silo:")
        print("-------------------------")
        clustering_results = results.get('clustering_results', {})
        
        # Export to Excel
        with pd.ExcelWriter("CLUSTER_STATISTICS.xlsx", engine='openpyxl') as writer:
            for silo_name, cluster_stats in clustering_results['cluster_stats'].items():
                # Create a DataFrame with all features
                data = []
                for cluster_id, stats in cluster_stats.items():
                    row_data = {'Cluster': cluster_id}
                    for feature in config['features']:
                        # Replace underscore with space in feature names
                        feature_name = feature.replace('_', ' ')
                        row_data.update({
                            f'{feature_name}_mean': round(stats['mean'][feature], 4),
                            f'{feature_name}_max': round(stats['max'][feature], 4),
                            f'{feature_name}_min': round(stats['min'][feature], 4),
                            f'{feature_name}_std': round(stats['std'][feature], 6)
                        })
                    data.append(row_data)
                
                # Create DataFrame
                df = pd.DataFrame(data).set_index('Cluster')
                
                # Create multi-index columns
                columns = pd.MultiIndex.from_product([
                    [f.replace('_', ' ') for f in config['features']], 
                    ['mean', 'max', 'min', 'std']
                ])
                
                # Reorder the DataFrame columns to match the multi-index
                reordered_cols = []
                for feature in config['features']:
                    feature_name = feature.replace('_', ' ')
                    reordered_cols.extend([
                        f'{feature_name}_mean',
                        f'{feature_name}_max',
                        f'{feature_name}_min',
                        f'{feature_name}_std'
                    ])
                df = df[reordered_cols]
                df.columns = columns
                
                sheet_name = f"{silo_name}_Statistics"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                
                # Write to Excel with the exact format as cluster_groups_agglo.csv
                df.to_excel(writer, sheet_name=sheet_name)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Adjust column widths
                from openpyxl.utils import get_column_letter
                for idx in range(1, len(df.columns) + 2):
                    col_letter = get_column_letter(idx)
                    max_length = 0
                    for row in range(1, len(df) + 3):
                        cell = worksheet.cell(row=row, column=idx)
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
        
        print("\nCluster statistics have been exported to 'cluster_statistics.xlsx'")
        
        for silo_name, summary in clustering_results['summary'].items():
            print(f"\nSILO: {silo_name}")
            print(f"Number of clusters: {summary['n_clusters']}")
            print(f"Number of samples: {summary['n_samples']}")
            print(f"Silhouette Score: {summary['silhouette_score']:.3f}")
            print("\nCluster sizes:")
            for cluster_id, size in summary['cluster_sizes'].items():
                print(f"Cluster {cluster_id}: {size} samples")

    except Exception as e:
        print(f"Error running pipeline: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()